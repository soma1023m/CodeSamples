from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file

#To be able to include images (jpeg, png, bmp,…) into an openpyxl file, you will also need the “pillow” library that can be installed with:
#pip install pillow
#Refer https://openpyxl.readthedocs.io/en/stable/
wb.save("sample.xlsx")
